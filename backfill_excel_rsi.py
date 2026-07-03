import argparse
import os
import shutil
import time

import requests
import urllib3
from dotenv import load_dotenv

try:
    from tradingview_ta import TA_Handler, Interval as TVInterval
except ImportError:
    TA_Handler = None
    TVInterval = None

from openpyxl import load_workbook

load_dotenv()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

ALPHA_VANTAGE_KEY = os.environ.get("ALPHA_VANTAGE_KEY", "")
RSI_PROVIDER = os.environ.get("RSI_PROVIDER", "auto").strip().lower()
RSI_INTERVAL = os.environ.get("RSI_INTERVAL", "15min")
RSI_PERIOD = int(os.environ.get("RSI_PERIOD", "14"))
RSI_SERIES_TYPE = os.environ.get("RSI_SERIES_TYPE", "close")
RSI_SYMBOL_SUFFIXES = [
    item.strip()
    for item in os.environ.get("RSI_SYMBOL_SUFFIXES", "NSE,BSE,").split(",")
    if item is not None
]
RSI_TV_EXCHANGES = [
    item.strip()
    for item in os.environ.get("RSI_TV_EXCHANGES", "NSE,BSE").split(",")
    if item and item.strip()
]
RSI_CACHE_TTL_SECONDS = int(os.environ.get("RSI_CACHE_TTL_SECONDS", "300"))

TV_INTERVAL_MAP = {
    "1min": TVInterval.INTERVAL_1_MINUTE if TVInterval else None,
    "5min": TVInterval.INTERVAL_5_MINUTES if TVInterval else None,
    "15min": TVInterval.INTERVAL_15_MINUTES if TVInterval else None,
    "30min": TVInterval.INTERVAL_30_MINUTES if TVInterval else None,
    "60min": TVInterval.INTERVAL_1_HOUR if TVInterval else None,
    "1h": TVInterval.INTERVAL_1_HOUR if TVInterval else None,
    "4h": TVInterval.INTERVAL_4_HOURS if TVInterval else None,
    "1day": TVInterval.INTERVAL_1_DAY if TVInterval else None,
    "daily": TVInterval.INTERVAL_1_DAY if TVInterval else None,
    "1week": TVInterval.INTERVAL_1_WEEK if TVInterval else None,
    "weekly": TVInterval.INTERVAL_1_WEEK if TVInterval else None,
    "1month": TVInterval.INTERVAL_1_MONTH if TVInterval else None,
    "monthly": TVInterval.INTERVAL_1_MONTH if TVInterval else None,
}

RSI_CACHE = {}


def get_rsi_action(rsi_value: float) -> str:
    if rsi_value < 30:
        return "Hold"
    if rsi_value <= 70:
        return "Buy"
    return "Sell"


def _get_cached_rsi(symbol: str):
    cached = RSI_CACHE.get(symbol)
    if not cached:
        return None

    age_seconds = time.time() - cached["timestamp"]
    if age_seconds > RSI_CACHE_TTL_SECONDS:
        return None

    return cached["rsi"], cached["action"]


def _set_cached_rsi(symbol: str, rsi_value: float, action: str) -> None:
    RSI_CACHE[symbol] = {
        "rsi": round(float(rsi_value), 2),
        "action": action,
        "timestamp": time.time(),
    }


def _get_last_cached_rsi(symbol: str):
    cached = RSI_CACHE.get(symbol)
    if not cached:
        return None
    return cached["rsi"], cached["action"]


def _resolve_tv_interval():
    return TV_INTERVAL_MAP.get(RSI_INTERVAL.lower()) or (
        TVInterval.INTERVAL_15_MINUTES if TVInterval else None
    )


def _fetch_rsi_from_tradingview(symbol: str):
    if TA_Handler is None:
        return None

    interval = _resolve_tv_interval()
    if interval is None:
        return None

    for exchange in RSI_TV_EXCHANGES:
        original_post = requests.post
        try:
            def _post_no_verify(*args, **kwargs):
                kwargs.setdefault("verify", False)
                return original_post(*args, **kwargs)

            requests.post = _post_no_verify
            handler = TA_Handler(
                symbol=symbol,
                screener="india",
                exchange=exchange,
                interval=interval,
            )
            analysis = handler.get_analysis()
            rsi_value = analysis.indicators.get("RSI")
            if rsi_value is None:
                continue

            action = get_rsi_action(float(rsi_value))
            return round(float(rsi_value), 2), action
        except Exception:
            continue
        finally:
            requests.post = original_post

    return None


def _fetch_rsi_from_alpha_vantage(symbol: str):
    if not ALPHA_VANTAGE_KEY:
        return None

    for suffix in RSI_SYMBOL_SUFFIXES:
        av_symbol = f"{symbol}.{suffix}" if suffix else symbol
        url = (
            "https://www.alphavantage.co/query"
            f"?function=RSI&symbol={av_symbol}"
            f"&interval={RSI_INTERVAL}&time_period={RSI_PERIOD}&series_type={RSI_SERIES_TYPE}"
            f"&apikey={ALPHA_VANTAGE_KEY}"
        )

        response = requests.get(url, timeout=20, verify=False)
        data = response.json()

        if "Note" in data or "Information" in data:
            return None

        if "Error Message" in data:
            continue

        rsi_data = data.get("Technical Analysis: RSI", {})
        if not rsi_data:
            continue

        latest_time = next(iter(rsi_data))
        latest_rsi = float(rsi_data[latest_time]["RSI"])
        action = get_rsi_action(latest_rsi)
        return round(latest_rsi, 2), action

    return None


def calculate_rsi(symbol: str):
    cached = _get_cached_rsi(symbol)
    if cached:
        return cached

    result = None
    if RSI_PROVIDER in ("auto", "tradingview"):
        result = _fetch_rsi_from_tradingview(symbol)
        if result:
            _set_cached_rsi(symbol, result[0], result[1])
            return result

    if RSI_PROVIDER in ("auto", "alpha_vantage"):
        result = _fetch_rsi_from_alpha_vantage(symbol)
        if result:
            _set_cached_rsi(symbol, result[0], result[1])
            return result

    stale_cached = _get_last_cached_rsi(symbol)
    if stale_cached:
        return stale_cached

    return 0.0, "No Data"


def find_column_indexes(header_row):
    normalized = {}
    for index, cell in enumerate(header_row, start=1):
        value = str(cell.value).strip().lower() if cell.value is not None else ""
        normalized[value] = index

    required = ["stocksymbol", "rsi", "action"]
    missing = [col for col in required if col not in normalized]
    if missing:
        raise ValueError(
            "Missing required columns in header row: " + ", ".join(missing)
        )

    return normalized["stocksymbol"], normalized["rsi"], normalized["action"]


def backfill_excel(file_path: str, sheet_name: str) -> None:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    backup_path = file_path + ".bak"
    shutil.copyfile(file_path, backup_path)

    workbook = load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")

    worksheet = workbook[sheet_name]

    header_row_index = 1
    stock_col, rsi_col, action_col = find_column_indexes(worksheet[header_row_index])

    updated_rows = 0
    skipped_rows = 0

    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        symbol_value = worksheet.cell(row=row_index, column=stock_col).value
        if symbol_value is None:
            skipped_rows += 1
            continue

        symbol = str(symbol_value).strip().upper()
        if not symbol:
            skipped_rows += 1
            continue

        rsi_value, action = calculate_rsi(symbol)
        worksheet.cell(row=row_index, column=rsi_col).value = float(rsi_value)
        worksheet.cell(row=row_index, column=action_col).value = action
        updated_rows += 1

        if updated_rows % 25 == 0:
            print(f"Updated {updated_rows} rows...")

    workbook.save(file_path)
    print(
        f"Backfill complete. Updated={updated_rows}, Skipped={skipped_rows}, Backup={backup_path}"
    )


def main():
    parser = argparse.ArgumentParser(
        description="Backfill RSI and Action columns for all rows in an Excel sheet."
    )
    parser.add_argument("--file", required=True, help="Path to the Excel file")
    parser.add_argument("--sheet", default="StockAlerts", help="Sheet name")
    args = parser.parse_args()

    backfill_excel(args.file, args.sheet)


if __name__ == "__main__":
    main()
